import logging
import secrets
import os
import re
import unicodedata
import requests
from datetime import datetime
from difflib import get_close_matches
from functools import wraps
from io import BytesIO
from werkzeug.utils import secure_filename
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    make_response
)
from models import (
    db,
    ConfiguracaoSistema,
    Banner,
    Carrossel,
    Aluno,
    Professor,
    Classe,
    Grupo,
    Turma,
    Comunicado,
    FAQ,
    EstatisticaPagina,
    PautaTurma,
    Sobre,
    Diretor,
    Contacto,
    PendenciaPauta,
    Nota,
    NotaTemporaria,
    Aviso,
    Publicacao
)

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None

controle_bp = Blueprint("controle", __name__)


# ==========================================
# 1. CLASSES FAKE / FALLBACK (POO)
# ==========================================

class BannerFake:
    def __init__(self):
        self.status = "normal"
        self.titulo = "Bem-vindo"
        self.mensagem = "Portal ESGAM"
        self.link = None
        self.link_texto = ""


class SobreFake:
    def __init__(self):
        self.eyebrow = "ESGAM"
        self.titulo = "Escola Secundária Geral de Alto Molócuè"
        self.texto = "Informações institucionais ainda não foram cadastradas."
        self.foto = "placeholder.png"


class DiretorFake:
    def __init__(self):
        self.titulo = "Mensagem da Direção"
        self.texto = "Mensagem ainda não disponível."
        self.foto = "placeholder.png"


class FAQFake:
    def __init__(self):
        self.id = ""
        self.pergunta = ""
        self.resposta = ""


class ContactoFake:
    def __init__(self):
        self.email = "esgam@email.com"
        self.telefone = "-"


# ==========================================
# 2. AUTENTICAÇÃO E SEGURANÇA
# ==========================================

def login_required(f):
    """Decorator para restrição de acesso a utilizadores autenticados.

    Além de bloquear o acesso, marca a resposta como "não guardável" (no-store).
    Isto evita que, depois de clicar em "Sair", o botão Voltar do navegador
    mostre uma versão em cache desta página protegida.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Por favor, efetue login para aceder a esta página.', 'erro')
            return redirect(url_for('login.login'))

        response = make_response(f(*args, **kwargs))
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    return decorated_function


# ==========================================
# 3. SERVIÇOS DA PÁGINA INICIAL (INDEX)
# ==========================================

def obter_banner():
    try:
        banner = Banner.query.filter_by(ativo=True).first()
        if banner:
            return banner
    except Exception:
        logging.exception("Erro ao carregar banner.")
        db.session.rollback()

    return BannerFake()


def obter_carrossel():
    try:
        return Carrossel.query.filter_by(ativo=True).order_by(Carrossel.ordem).all()
    except Exception:
        logging.exception("Erro ao carregar carrossel.")
        db.session.rollback()
        return []


def obter_estatisticas():
    def contar(modelo):
        try:
            return int(modelo.query.count())
        except Exception:
            db.session.rollback()
            return 0

    defaults = [
        ("ANO", "Ano de Fundação", 2000.0, False, ""),
        ("ALUNO", "Alunos Registados", contar(Aluno), True, "+"),
        ("PROF", "Corpo Docente", contar(Professor), True, ""),
        ("TURMA", "Turmas Activas", contar(Turma), True, ""),
    ]

    try:
        itens = {item.cod: item for item in EstatisticaPagina.query.order_by(EstatisticaPagina.id.asc()).all()}
        for cod, label, valor, animar, sufixo in defaults:
            if cod not in itens:
                db.session.add(EstatisticaPagina(cod=cod, label=label, valor=valor, animar=animar, sufixo=sufixo))
        db.session.commit()
        return EstatisticaPagina.query.order_by(EstatisticaPagina.id.asc()).all()
    except Exception:
        db.session.rollback()
        return [
            {"id": 1, "cod": "ANO", "label": "Ano de Fundação", "valor": 2000, "animar": False, "sufixo": ""},
            {"id": 2, "cod": "ALUNO", "label": "Alunos Registados", "valor": contar(Aluno), "animar": True, "sufixo": "+"},
            {"id": 3, "cod": "PROF", "label": "Corpo Docente", "valor": contar(Professor), "animar": True, "sufixo": ""},
            {"id": 4, "cod": "TURMA", "label": "Turmas Activas", "valor": contar(Turma), "animar": True, "sufixo": ""}
        ]

def obter_comunicados():
    try:
        return Comunicado.query.order_by(Comunicado.data.desc()).limit(1).all()
    except Exception:
        logging.exception("Erro ao carregar comunicados.")
        db.session.rollback()
        return []


def obter_faq():
    try:
        faq = FAQ.query.first()
        if faq:
            return [faq]
        faq = FAQ(pergunta='Como consultar as pautas do meu educando?', resposta='Aceda ao portal e use o seu ID e senha.')
        db.session.add(faq)
        db.session.commit()
        return [faq]
    except Exception:
        logging.exception("Erro ao carregar FAQ.")
        db.session.rollback()
        return [FAQFake()]


def obter_sobre():
    try:
        sobre = Sobre.query.first()
        if sobre:
            return sobre
    except Exception:
        logging.exception("Erro ao carregar Sobre.")
        db.session.rollback()

    return SobreFake()


def obter_diretor():
    try:
        diretor = Diretor.query.first()
        if diretor:
            return diretor
    except Exception:
        logging.exception("Erro ao carregar Diretor.")
        db.session.rollback()

    return DiretorFake()


def obter_contacto():
    try:
        contacto = Contacto.query.first()
        if contacto:
            return contacto
    except Exception:
        logging.exception("Erro ao carregar contacto.")
        db.session.rollback()

    return ContactoFake()


def carregar_index():
    """Agrupa todos os dados necessários para renderizar a página inicial."""
    return {
        "banner": obter_banner(),
        "fotos_carrossel": obter_carrossel(),
        "estatisticas": obter_estatisticas(),
        "comunicados": obter_comunicados(),
        "faq_dinamico": obter_faq(),
        "sobre": obter_sobre(),
        "diretor": obter_diretor(),
        "contacto": obter_contacto(),
        "ano_atual": datetime.now().year
    }


# ==========================================
# 4. GESTÃO DE PAUTAS E VERIFICAÇÃO DE DADOS
# ==========================================

@controle_bp.route('/admin/verificacao', methods=['GET'])
@login_required
def central_verificacao():
    pendencias = PendenciaPauta.query.filter_by(status='pendente').order_by(PendenciaPauta.id.desc()).all()

    config = ConfiguracaoSistema.query.first()

    faqs = obter_faq()
    faq_unica = faqs[0] if faqs else FAQFake()
    contexto = {
        "total_alunos": Aluno.query.filter_by(deleted_at=None).count(),
        "total_classes": Classe.query.filter_by(deleted_at=None).count(),
        "total_grupos": Grupo.query.filter_by(deleted_at=None).count(),
        "trimestre_atual": (config.ano_letivo if config else datetime.now().year),
        "banner": obter_banner(),
        "sobre": obter_sobre(),
        "diretor": obter_diretor(),
        "contacto": obter_contacto(),
        "ano_atual": datetime.now().year,
        "estatisticas": obter_estatisticas(),
        "fotos_carrossel": obter_carrossel(),
        "faq_dinamico": faqs,
        "faq_unica": faq_unica,
        "publicacoes": Publicacao.query.filter_by(ativo=True).order_by(Publicacao.data_publicacao.desc()).all(),
        "excel_importados": Nota.query.count(),
        "excel_recebidos": Nota.query.count() + PendenciaPauta.query.filter_by(status='pendente').count(),
        "pendencias": pendencias,
    }

    return render_template('controle.html', **contexto)


@controle_bp.route('/admin/publicacoes', methods=['GET', 'POST'])
@login_required
def gerir_publicacoes():
    if request.method == 'POST':
        titulo = request.form.get('titulo', '').strip()
        titulo = titulo[:100]
        categoria = request.form.get('categoria', 'Outro').strip()
        descricao = request.form.get('descricao', '').strip()
        classe = request.form.get('classe', '').strip()
        arquivo = request.files.get('arquivo')

        if not titulo or not arquivo or not arquivo.filename:
            flash('Título e PDF são obrigatórios.', 'erro')
            return redirect(url_for('controle.central_verificacao'))

        url_arquivo = _guardar_ficheiro('arquivo', subpasta='publicacoes')
        if not url_arquivo:
            return redirect(url_for('controle.central_verificacao'))

        publicacao = Publicacao(
            titulo=titulo,
            categoria=categoria or 'Outro',
            descricao=descricao or None,
            classe=classe or None,
            arquivo=url_arquivo,
            ativo=True
        )
        db.session.add(publicacao)
        db.session.commit()
        flash('Publicação criada com sucesso.', 'sucesso')

    return redirect(url_for('controle.central_verificacao'))


SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://gmiwrafjeqpixesqvuxe.supabase.co").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
SUPABASE_BUCKET = "uploads"


def _guardar_bytes_supabase(nome_arquivo, conteudo, tipo_mime, subpasta=''):
    """Envia bytes para o Supabase Storage e devolve o URL público, ou None.

    Não toca no disco local — funciona igual em local e no Vercel (cujo
    filesystem é só-de-leitura fora de /tmp).
    """
    if not SUPABASE_KEY:
        logging.error("SUPABASE_KEY não está definida — não é possível enviar ficheiros.")
        flash('Configuração de armazenamento em falta. Contacte o suporte técnico.', 'erro')
        return None

    nome_arquivo = secure_filename(nome_arquivo)
    caminho_storage = f"{subpasta}/{nome_arquivo}" if subpasta else nome_arquivo

    url_upload = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{caminho_storage}"
    headers = {
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "apikey": SUPABASE_KEY,
        "Content-Type": tipo_mime or 'application/octet-stream',
        "x-upsert": "true",  # substitui o ficheiro se já existir com o mesmo nome
    }

    try:
        resposta = requests.put(url_upload, headers=headers, data=conteudo, timeout=30)
        if resposta.status_code not in (200, 201):
            logging.error("Erro ao enviar '%s' para o Supabase Storage: %s - %s",
                          nome_arquivo, resposta.status_code, resposta.text)
            return None
    except requests.RequestException:
        logging.exception("Erro de rede ao enviar '%s' para o Supabase Storage.", nome_arquivo)
        return None

    return f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{caminho_storage}"


def _guardar_ficheiro(campo_nome, subpasta=''):
    """Envia um ficheiro de imagem (vindo de request.files) para o Supabase Storage."""
    arquivo = request.files.get(campo_nome)
    if not (arquivo and arquivo.filename):
        return None

    nome_arquivo = secure_filename(arquivo.filename)
    conteudo = arquivo.read()
    tipo_mime = arquivo.mimetype or 'application/octet-stream'

    url_publico = _guardar_bytes_supabase(nome_arquivo, conteudo, tipo_mime, subpasta=subpasta)
    if not url_publico:
        flash('Não foi possível guardar a imagem. As restantes alterações foram guardadas.', 'erro')
    return url_publico


def _extension_excel(nome_arquivo):
    return os.path.splitext(nome_arquivo.lower())[1] in {".xlsx", ".xls"}


def _subpasta_para_tipo_pauta(tipo_pauta):
    tipo = (tipo_pauta or "").strip().lower()
    if tipo == "notas":
        return "controle"
    if tipo == "turmas":
        return "pautas"
    return "pautas"


# ------------------------------------------------------------------
# Leitura e classificação das linhas do Excel de pautas
# ------------------------------------------------------------------

CABECALHOS_ESPERADOS = {
    "matricula": ("matricula", "matrícula", "codigo", "código", "id"),
    "nome": ("nome", "aluno", "nome do aluno", "estudante"),
    "classe": ("classe",),
    "turma": ("turma",),
    "periodo": ("periodo", "período", "trimestre"),
    "disciplina": ("disciplina",),
    "nota_ac": ("ac", "nota_ac", "notaac", "1p", "p1"),
    "nota_pt": ("pt", "nota_pt", "notapt", "2p", "p2"),
    "nota_ap": ("ap", "nota_ap", "notaap", "3p", "p3", "4p", "p4"),
    "nota_exame": ("x", "exame", "nota_x", "notax", "exam", "nota exame"),
}


def _normalizar(texto):
    """Remove acentos e baixa a caixa, para comparar cabeçalhos/nomes com tolerância."""
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", texto).strip()


def _mapear_colunas(linha_cabecalho):
    """Associa cada coluna do Excel (por uma linha de cabeçalho) ao campo interno correspondente."""
    mapa = {}
    for indice, celula in enumerate(linha_cabecalho or []):
        chave_normalizada = _normalizar(celula)
        for campo, aliases in CABECALHOS_ESPERADOS.items():
            if chave_normalizada in aliases:
                mapa[campo] = indice
                break
    return mapa


def _localizar_linha_cabecalho(linhas_brutas, limite=15):
    """Procura, nas primeiras linhas do ficheiro, a linha que contém a coluna 'Nome'."""
    for indice, linha in enumerate(linhas_brutas[:limite]):
        for celula in (linha or []):
            if _normalizar(celula) in CABECALHOS_ESPERADOS["nome"]:
                return indice
    return None


def _extrair_classe_turma_do_titulo(linhas_brutas, limite=10):
    """Título tipo 'PAUTA...11ª Classe GRUPO: A TURMA: B' → (classe, turma, grupo)."""
    for linha in linhas_brutas[:limite]:
        texto = " ".join(str(c) for c in (linha or []) if c is not None)
        if not texto:
            continue
        norm = _normalizar(texto)
        if "classe" not in norm and "grupo" not in norm and "turma" not in norm:
            continue

        classe = turma = grupo = None

        m_classe = re.search(r"(\d+)\s*ª?\s*classe", texto, re.IGNORECASE)
        if m_classe:
            classe = m_classe.group(1)

        m_turma = re.search(r"turma[:\s]+([A-Za-z0-9]+)", texto, re.IGNORECASE)
        if m_turma:
            turma = m_turma.group(1).upper()

        m_grupo = re.search(r"grupo[:\s]+([A-Za-z0-9]+)", texto, re.IGNORECASE)
        if m_grupo:
            grupo = m_grupo.group(1).upper()

        if classe or turma or grupo:
            return classe, turma, grupo

    return None, None, None


def _mapear_disciplinas_formato_largo(linha_cabecalho, coluna_nome):
    """
    Formato oficial de pauta: várias disciplinas lado a lado na mesma linha de
    cabeçalho, cada uma ocupando várias colunas (normalmente 3 trimestres + média).
    Devolve uma lista de (nome_disciplina, coluna_inicial, largura), ignorando a
    última coluna se for "Resultado" (não é uma disciplina).
    """
    posicoes = []
    for indice, celula in enumerate(linha_cabecalho or []):
        if indice <= coluna_nome or celula is None:
            continue
        texto = str(celula).strip()
        if not texto:
            continue
        if _normalizar(texto) in ("resultado", "genero", "género", "sexo"):
            continue
        posicoes.append((texto, indice))

    disciplinas = []
    for i, (nome_disciplina, col_inicio) in enumerate(posicoes):
        col_fim = posicoes[i + 1][1] if i + 1 < len(posicoes) else col_inicio + 4
        disciplinas.append((nome_disciplina, col_inicio, col_fim - col_inicio))

    return disciplinas


def _ler_linhas_excel(conteudo_bytes):
    """Lê o ficheiro Excel (em memória) e devolve (linhas, erro).

    Reconhece dois formatos:
    1. Simples — 1ª linha com cabeçalhos: Matrícula | Nome | Classe | Turma |
       Período | Disciplina | AC | PT | AP (uma disciplina por linha).
    2. Pauta oficial — cabeçalho pode estar em qualquer linha inicial, com várias
       disciplinas lado a lado (cada uma com 1º/2º/3º trimestre + média). Classe e
       turma são detetados a partir do título da pauta (ex: "...11ª Classe GRUPO:
       A TURMA: B..."). Os 3 trimestres de cada disciplina são guardados em
       nota_ac, nota_pt e nota_ap.
    """
    if load_workbook is None:
        return None, "A biblioteca 'openpyxl' não está instalada no servidor — contacte o suporte técnico."

    try:
        livro = load_workbook(BytesIO(conteudo_bytes), data_only=True, read_only=True)
        folha = livro.active
    except Exception:
        logging.exception("Erro ao abrir o ficheiro Excel.")
        return None, "O ficheiro não pôde ser aberto. Verifique se é um .xlsx válido."

    linhas_brutas = list(folha.iter_rows(values_only=True))
    if not linhas_brutas:
        return None, "O ficheiro Excel está vazio."

    indice_cabecalho = _localizar_linha_cabecalho(linhas_brutas)
    if indice_cabecalho is None:
        return None, ("Não foi possível encontrar uma coluna 'Nome' em nenhuma das primeiras "
                       "linhas do ficheiro. Confirme que o Excel tem essa coluna.")

    linha_cabecalho = linhas_brutas[indice_cabecalho]
    mapa = _mapear_colunas(linha_cabecalho)

    def para_float(v):
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    linhas = []

    if "disciplina" in mapa:
        # ---- Formato simples: uma disciplina por linha ----
        def valor(linha, campo):
            indice = mapa.get(campo)
            if indice is None or indice >= len(linha):
                return None
            return linha[indice]

        for linha in linhas_brutas[indice_cabecalho + 1:]:
            nome = valor(linha, "nome")
            if nome is None or str(nome).strip() == "":
                continue

            linhas.append({
                "matricula": (str(valor(linha, "matricula")).strip() if valor(linha, "matricula") not in (None, "") else None),
                "nome": str(nome).strip(),
                "classe": str(valor(linha, "classe") or "").strip(),
                "turma": str(valor(linha, "turma") or "").strip(),
                "periodo": str(valor(linha, "periodo") or "").strip(),
                "disciplina": str(valor(linha, "disciplina") or "").strip(),
                "nota_ac": para_float(valor(linha, "nota_ac")),
                "nota_pt": para_float(valor(linha, "nota_pt")),
                "nota_ap": para_float(valor(linha, "nota_ap")),
                "nota_exame": para_float(valor(linha, "nota_exame")),
            })
    else:
        # ---- Formato oficial de pauta: disciplinas lado a lado ----
        coluna_nome = mapa["nome"]
        disciplinas = _mapear_disciplinas_formato_largo(linha_cabecalho, coluna_nome)
        if not disciplinas:
            return None, ("Encontrei a coluna 'Nome', mas não consegui identificar nenhuma "
                           "disciplina no cabeçalho. Confirme o formato do ficheiro.")

        classe_detetada, turma_detetada, grupo_detetado = _extrair_classe_turma_do_titulo(linhas_brutas[:indice_cabecalho])

        for linha in linhas_brutas[indice_cabecalho + 1:]:
            nome = linha[coluna_nome] if coluna_nome < len(linha) else None
            if nome is None or str(nome).strip() == "":
                continue  # ignora linhas de subcabeçalho e espaços em branco preparados

            nome = str(nome).strip()

            for nome_disciplina, col_inicio, largura in disciplinas:
                valores = [
                    linha[col_inicio + i] if (col_inicio + i) < len(linha) else None
                    for i in range(min(largura, 3))  # só os 3 trimestres, ignora a média
                ]
                if all(v in (None, "") for v in valores):
                    continue  # disciplina sem nenhuma nota lançada para este aluno

                linhas.append({
                    "matricula": None,
                    "nome": nome,
                    "classe": classe_detetada or "",
                    "turma": turma_detetada or "",
                    "grupo": grupo_detetado or "",
                    "periodo": "",
                    "disciplina": str(nome_disciplina).strip(),
                    "nota_ac": para_float(valores[0] if len(valores) > 0 else None),
                    "nota_pt": para_float(valores[1] if len(valores) > 1 else None),
                    "nota_ap": para_float(valores[2] if len(valores) > 2 else None),
                    "nota_exame": None,
                })

    if not linhas:
        return None, "Nenhuma linha com nome de aluno e notas foi encontrada no ficheiro."

    return linhas, None



def _localizar_aluno(matricula, nome, classe):
    """Tenta encontrar o aluno correspondente. Devolve (aluno_exato, aluno_semelhante)."""
    if matricula:
        aluno = Aluno.query.filter_by(codigo_estudante=matricula, deleted_at=None).first()
        if aluno:
            return aluno, None

    candidatos_query = Aluno.query.filter_by(deleted_at=None)
    if classe:
        classe_obj = Classe.query.filter_by(nome=classe, deleted_at=None).first() \
            or (Classe.query.filter_by(numero=int(classe), deleted_at=None).first() if classe.isdigit() else None)
        if classe_obj:
            candidatos_query = candidatos_query.filter_by(classe_id=classe_obj.id)
    candidatos = candidatos_query.all()

    nome_normalizado = _normalizar(nome)
    for candidato in candidatos:
        if _normalizar(candidato.nome) == nome_normalizado:
            return candidato, None

    nomes_candidatos = {c.nome: c for c in candidatos}
    semelhantes = get_close_matches(nome, list(nomes_candidatos.keys()), n=1, cutoff=0.75)
    if semelhantes:
        return None, nomes_candidatos[semelhantes[0]]

    return None, None



def _gerar_codigo_estudante_unico():
    """Gera ID único e difícil de adivinhar (ex: ESG-A3F9K2)."""
    for _ in range(50):
        codigo = "ESG-" + secrets.token_hex(3).upper()
        if not Aluno.query.filter_by(codigo_estudante=codigo).first():
            return codigo
    return "ESG-" + secrets.token_hex(8).upper()


def _registar_pendencia(tipo, arquivo_nome, dados, nome_banco=None, descricao="", aluno_id=None):
    """Cria pendência só se ainda não existir uma igual (evita spam de duplicados)."""
    q = PendenciaPauta.query.filter_by(
        status='pendente',
        tipo=tipo,
        nome_excel=dados.get("nome"),
        classe=dados.get("classe"),
        turma=dados.get("turma"),
        periodo=dados.get("periodo"),
    )
    existente = q.first()
    if existente:
        # Atualiza descrição/arquivo se já houver pendência igual
        existente.arquivo = arquivo_nome or existente.arquivo
        existente.descricao = descricao or existente.descricao
        existente.nome_banco = nome_banco or existente.nome_banco
        # Atualiza notas temporárias da mesma disciplina, se existirem
        nt = NotaTemporaria.query.filter_by(
            pendencia_id=existente.id,
            disciplina=dados.get("disciplina"),
        ).first()
        if nt:
            nt.nota_ac = dados.get("nota_ac")
            nt.nota_pt = dados.get("nota_pt")
            nt.nota_ap = dados.get("nota_ap")
            nt.nota_exame = dados.get("nota_exame")
            nt.aluno_id = aluno_id if aluno_id is not None else nt.aluno_id
        else:
            db.session.add(NotaTemporaria(
                pendencia_id=existente.id,
                aluno_id=aluno_id,
                disciplina=dados.get("disciplina"),
                classe=dados.get("classe"),
                turma=dados.get("turma"),
                periodo=dados.get("periodo"),
                nota_ac=dados.get("nota_ac"),
                nota_pt=dados.get("nota_pt"),
                nota_ap=dados.get("nota_ap"),
                nota_exame=dados.get("nota_exame"),
            ))
        return existente

    pendencia = PendenciaPauta(
        arquivo=arquivo_nome,
        classe=dados.get("classe"),
        turma=dados.get("turma"),
        periodo=dados.get("periodo"),
        tipo=tipo,
        descricao=descricao,
        nome_excel=dados.get("nome"),
        nome_banco=nome_banco,
        status='pendente',
    )
    db.session.add(pendencia)
    db.session.flush()

    db.session.add(NotaTemporaria(
        pendencia_id=pendencia.id,
        aluno_id=aluno_id,
        disciplina=dados.get("disciplina"),
        classe=dados.get("classe"),
        turma=dados.get("turma"),
        periodo=dados.get("periodo"),
        nota_ac=dados.get("nota_ac"),
        nota_pt=dados.get("nota_pt"),
        nota_ap=dados.get("nota_ap"),
        nota_exame=dados.get("nota_exame"),
    ))
    return pendencia


def _turma_ja_tem_notas(classe, turma, periodo):
    """True se já existem notas para esta turma + período."""
    q = Nota.query
    if classe:
        q = q.filter_by(classe=classe)
    if turma:
        q = q.filter_by(turma=turma)
    if periodo:
        q = q.filter_by(periodo=periodo)
    return q.first() is not None


def _garantir_hierarquia(classe_str, turma_str=None, grupo_str=None):
    """Cria/obtém Classe, Grupo e Turma. Devolve (classe_id, grupo_id, turma_id)."""
    classe_id = grupo_id = turma_id = None
    if not classe_str:
        return None, None, None

    classe_str = str(classe_str).strip()
    classe_obj = None
    if classe_str.isdigit():
        n = int(classe_str)
        classe_obj = Classe.query.filter_by(numero=n, deleted_at=None).first()
        if not classe_obj:
            classe_obj = Classe(numero=n, nome=f"{n}ª Classe")
            db.session.add(classe_obj)
            db.session.flush()
    if not classe_obj:
        classe_obj = Classe.query.filter(
            Classe.deleted_at.is_(None),
            Classe.nome.ilike(f"%{classe_str}%"),
        ).first()
    if classe_obj:
        classe_id = classe_obj.id

    if classe_id and grupo_str:
        g = str(grupo_str).strip().upper()
        grupo_obj = Grupo.query.filter_by(nome=g, classe_id=classe_id, deleted_at=None).first()
        if not grupo_obj:
            grupo_obj = Grupo(nome=g, classe_id=classe_id)
            db.session.add(grupo_obj)
            db.session.flush()
        grupo_id = grupo_obj.id

    if classe_id and turma_str:
        t = str(turma_str).strip().upper()
        q = Turma.query.filter_by(nome=t, classe_id=classe_id, deleted_at=None)
        if grupo_id:
            q = q.filter_by(grupo_id=grupo_id)
        turma_obj = q.first()
        if not turma_obj:
            turma_obj = Turma(nome=t, classe_id=classe_id, grupo_id=grupo_id)
            db.session.add(turma_obj)
            db.session.flush()
        turma_id = turma_obj.id

    return classe_id, grupo_id, turma_id


def _aplicar_hierarquia_aluno(aluno, dados):
    c_id, g_id, t_id = _garantir_hierarquia(
        dados.get("classe"), dados.get("turma"), dados.get("grupo")
    )
    if c_id:
        aluno.classe_id = c_id
    if g_id:
        aluno.grupo_id = g_id
    if t_id:
        aluno.turma_id = t_id


def _substituir_notas_turma(classe, turma):
    """Apaga notas antigas da mesma classe+turma (republicação)."""
    if not classe and not turma:
        return 0
    q = Nota.query
    if classe:
        q = q.filter(Nota.classe.ilike(f"%{str(classe).strip()}%"))
    if turma:
        q = q.filter(Nota.turma.ilike(f"%{str(turma).strip()}%"))
    n = q.count()
    q.delete(synchronize_session=False)
    return n


def _criar_aluno_automatico(dados):
    """Cria aluno ESG e liga classe/grupo/turma da pauta."""
    codigo = _gerar_codigo_estudante_unico()
    aluno = Aluno(nome=dados["nome"], codigo_estudante=codigo)
    _aplicar_hierarquia_aluno(aluno, dados)
    db.session.add(aluno)
    db.session.flush()
    return aluno


def _guardar_nota(aluno_id, dados):
    """Guarda AC/PT/AP + exame (X) no model."""
    db.session.add(Nota(
        aluno_id=aluno_id,
        disciplina=dados["disciplina"],
        classe=dados.get("classe"),
        turma=dados.get("turma"),
        periodo=dados.get("periodo"),
        nota_ac=dados.get("nota_ac"),
        nota_pt=dados.get("nota_pt"),
        nota_ap=dados.get("nota_ap"),
        nota_exame=dados.get("nota_exame"),
    ))


def _processar_linha_pauta(dados, arquivo_nome, modo_substituicao=False):
    """
    Importa uma linha da pauta.
    modo_substituicao=True: notas da turma já foram limpas — grava sem pendência de duplicado.
    """
    if not dados.get("disciplina"):
        return "ignorado"

    aluno_exato, aluno_semelhante = _localizar_aluno(
        dados.get("matricula"), dados.get("nome"), dados.get("classe")
    )

    if aluno_exato:
        _aplicar_hierarquia_aluno(aluno_exato, dados)
        q = Nota.query.filter_by(aluno_id=aluno_exato.id, disciplina=dados["disciplina"])
        if dados.get("periodo"):
            q = q.filter_by(periodo=dados.get("periodo"))
        if dados.get("classe"):
            q = q.filter(Nota.classe.ilike(f"%{dados.get('classe')}%"))
        if dados.get("turma"):
            q = q.filter(Nota.turma.ilike(f"%{dados.get('turma')}%"))
        existentes = q.all()
        for n in existentes:
            db.session.delete(n)
        _guardar_nota(aluno_exato.id, dados)
        return "substituido" if existentes else "importado"

    if aluno_semelhante and not modo_substituicao:
        _registar_pendencia(
            "nome", arquivo_nome, dados,
            nome_banco=aluno_semelhante.nome,
            descricao=(
                f'O nome "{dados["nome"]}" não corresponde exatamente. '
                f'Encontrámos "{aluno_semelhante.nome}". Confirme se é o mesmo.'
            ),
        )
        return "nome"

    aluno = _criar_aluno_automatico(dados)
    _guardar_nota(aluno.id, dados)
    return "novo"



@controle_bp.route('/admin/banner', methods=['POST'])
@login_required
def atualizar_banner():
    status = request.form.get('status', 'info').strip()
    titulo = request.form.get('titulo', '').strip()
    mensagem = request.form.get('mensagem', '').strip()
    link_texto = request.form.get('link_texto', '').strip()
    titulo = titulo[:15]
    mensagem = mensagem[:70]

    banner = Banner.query.filter_by(ativo=True).first()
    if not banner:
        banner = Banner(titulo=titulo, mensagem=mensagem, ativo=True)
        db.session.add(banner)

    banner.status = status
    banner.titulo = titulo
    banner.mensagem = mensagem
    banner.link_texto = link_texto

    try:
        db.session.commit()
        flash('Banner atualizado com sucesso!', 'banner')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar o banner: {str(e)}', 'banner')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/sobre', methods=['POST'])
@login_required
def atualizar_sobre():
    texto = request.form.get('texto', '').strip()
    texto = texto[:260]

    sobre = Sobre.query.first()
    if not sobre:
        sobre = Sobre(titulo='Sobre a Escola', texto=texto)
        db.session.add(sobre)

    sobre.texto = texto

    nome_foto = _guardar_ficheiro('foto')
    if nome_foto:
        sobre.foto = nome_foto

    try:
        db.session.commit()
        flash('Informações atualizadas com sucesso!', 'sobre')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar informações: {str(e)}', 'sobre')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/diretor', methods=['POST'])
@login_required
def atualizar_diretor():
    titulo = "Mensagem da Direção"
    texto = request.form.get('texto', '').strip()[:500]

    diretor = Diretor.query.first()
    if not diretor:
        diretor = Diretor(titulo=titulo, texto=texto or "")
        db.session.add(diretor)
    else:
        diretor.titulo = titulo
        diretor.texto = texto

    try:
        nome_foto = _guardar_ficheiro('foto')
        if nome_foto:
            diretor.foto = nome_foto
    except Exception:
        logging.exception("Falha ao enviar fotografia do diretor")
        flash('A mensagem foi guardada, mas a fotografia não pôde ser enviada.', 'diretor')

    try:
        db.session.commit()
        flash('Mensagem da Direção atualizada com sucesso!', 'diretor')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar mensagem: {str(e)}', 'diretor')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/comunicado', methods=['POST'])
@login_required
def publicar_comunicado():
    titulo = request.form.get('titulo', '').strip()
    texto = request.form.get('texto', '').strip()
    titulo = titulo[:15]
    texto = texto[:180]

    if not titulo or not texto:
        flash('Título e mensagem são obrigatórios.', 'comunicado')
        return redirect(url_for('controle.central_verificacao'))

    novo = Comunicado(titulo=titulo, texto=texto)
    db.session.add(novo)

    try:
        db.session.commit()

        # Mantém apenas os 3 comunicados mais recentes
        todos = Comunicado.query.order_by(Comunicado.data.desc()).all()
        for antigo in todos[3:]:
            db.session.delete(antigo)
        db.session.commit()

        flash('Comunicado publicado com sucesso!', 'comunicado')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao publicar comunicado: {str(e)}', 'comunicado')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/contacto', methods=['POST'])
@login_required
def atualizar_contacto():
    email = request.form.get('email', '').strip()
    telefone = request.form.get('telefone', '').strip()
    telefone = ''.join(ch for ch in telefone if ch.isdigit())[:9]

    contacto = Contacto.query.first()
    if not contacto:
        contacto = Contacto(email=email or 'esgam@email.com', telefone=telefone or '-')
        db.session.add(contacto)

    if email:
        contacto.email = email
    if telefone:
        contacto.telefone = telefone

    try:
        db.session.commit()
        flash('Contactos atualizados com sucesso!', 'contacto')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar contactos: {str(e)}', 'contacto')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/estatisticas', methods=['POST'])
@login_required
def atualizar_estatisticas():
    try:
        ids = request.form.getlist('id')
        labels = request.form.getlist('label')
        valores = request.form.getlist('valor')
        sufixos = request.form.getlist('sufixo')

        if not ids:
            flash('Nenhuma estat?stica foi recebida.', 'estatisticas')
            return redirect(url_for('controle.central_verificacao'))

        itens = EstatisticaPagina.query.order_by(EstatisticaPagina.id.asc()).all()
        itens_por_id = {str(item.id): item for item in itens}
        for idx, stat_id in enumerate(ids):
            item = itens_por_id.get(str(stat_id))
            if not item:
                continue
            if idx < len(labels):
                item.label = labels[idx].strip()[:120] or item.label
            if idx < len(valores):
                try:
                    item.valor = int(float(valores[idx]))
                except (TypeError, ValueError):
                    pass
            if idx < len(sufixos):
                item.sufixo = sufixos[idx].strip()[:20]

        db.session.commit()
        flash('Estatísticas guardadas com sucesso!', 'estatisticas')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao guardar estatísticas: {str(e)}', 'estatisticas')

    return redirect(url_for('controle.central_verificacao'))

@controle_bp.route('/admin/faq', methods=['POST'])
@login_required
def atualizar_faq():
    faq_id = request.form.get('id')
    pergunta = request.form.get('pergunta', '').strip()
    resposta = request.form.get('resposta', '').strip()
    pergunta = pergunta[:40]
    resposta = resposta[:120]

    if not faq_id:
        flash('Selecione uma pergunta válida.', 'faq')
        return redirect(url_for('controle.central_verificacao'))

    item = FAQ.query.get(faq_id)
    if not item:
        item = FAQ(pergunta=pergunta, resposta=resposta)
        db.session.add(item)
    else:
        item.pergunta = pergunta
        item.resposta = resposta

    try:
        db.session.commit()
        flash('FAQ atualizada com sucesso!', 'faq')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar FAQ: {str(e)}', 'faq')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/publicar-aviso', methods=['POST'])
@login_required
def publicar_aviso():
    destino = request.form.get('destino', '').strip()
    mensagem = request.form.get('mensagem', '').strip()

    if not destino or not mensagem:
        flash('Destino e mensagem são obrigatórios.', 'aviso')
        return redirect(url_for('controle.central_verificacao'))

    # Desativa avisos anteriores e publica o novo como ativo
    Aviso.query.filter_by(ativo=True).update({'ativo': False})

    novo_aviso = Aviso(
        mensagem=f"[{destino}] {mensagem}",
        texto=mensagem,
        ativo=True
    )
    db.session.add(novo_aviso)

    try:
        db.session.commit()
        flash('Aviso publicado com sucesso!', 'aviso')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao publicar aviso: {str(e)}', 'aviso')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/pautas', methods=['GET'])
@login_required
def listar_pautas():
    pautas = PautaTurma.query.filter_by(ativo=True).order_by(PautaTurma.data_publicacao.desc()).all()
    return render_template('pauta.html', publicacoes=pautas, modo_pauta_aberto=True)

@controle_bp.route('/admin/upload-pauta', methods=['POST'])
@login_required
def upload_pauta():
    tipo_pauta = request.form.get('tipo_pauta', '').strip() or 'turmas'
    ficheiros = [f for f in request.files.getlist('arquivo_pauta') if f and f.filename]

    if not ficheiros:
        flash('Seleciona pelo menos um ficheiro Excel.', 'pauta')
        return redirect(url_for('controle.central_verificacao'))

    total_processados = 0
    mensagens_erro = []

    for arquivo in ficheiros:
        nome_original = secure_filename(arquivo.filename)
        if not _extension_excel(nome_original):
            mensagens_erro.append(f'"{arquivo.filename}": apenas ficheiros Excel (.xlsx ou .xls) são aceites.')
            continue

        try:
            conteudo = arquivo.read()
            pauta = PautaTurma(
                titulo=os.path.splitext(nome_original)[0],
                tipo=tipo_pauta,
                categoria='Turmas' if tipo_pauta == 'turmas' else 'Notas',
                classe=None,
                turma=None,
                periodo=None,
                arquivo=nome_original,
                ficheiro=conteudo,
                mimetype=arquivo.mimetype or 'application/octet-stream',
                ativo=True,
            )
            db.session.add(pauta)
            db.session.commit()
            total_processados += 1

            # Processamento inteligente só para notas
            if tipo_pauta == "notas":
                linhas, erro_leitura = _ler_linhas_excel(conteudo)
                if erro_leitura:
                    mensagens_erro.append(f'"{arquivo.filename}": {erro_leitura}')
                else:
                    amostra = next((d for d in linhas if d.get("classe") or d.get("turma")), {})
                    classe_p = (amostra.get("classe") or "").strip()
                    turma_p = (amostra.get("turma") or "").strip()
                    grupo_p = (amostra.get("grupo") or "").strip()

                    modo_sub = False
                    removidas = 0
                    if classe_p or turma_p:
                        if _turma_ja_tem_notas(classe_p, turma_p, None):
                            removidas = _substituir_notas_turma(classe_p, turma_p)
                            modo_sub = True

                    _garantir_hierarquia(classe_p, turma_p, grupo_p)

                    contadores = {
                        "importado": 0, "substituido": 0, "duplicado": 0,
                        "nome": 0, "novo": 0, "ignorado": 0,
                    }
                    for dados in linhas:
                        if grupo_p and not dados.get("grupo"):
                            dados["grupo"] = grupo_p
                        resultado = _processar_linha_pauta(
                            dados, nome_original, modo_substituicao=modo_sub
                        )
                        contadores[resultado] = contadores.get(resultado, 0) + 1
                    try:
                        db.session.commit()
                        if modo_sub:
                            flash(
                                f'"{arquivo.filename}": pauta {classe_p or "?"}ª turma {turma_p or "?"} '
                                f'substituída ({removidas} notas antigas; '
                                f'{contadores["importado"] + contadores["substituido"]} gravadas; '
                                f'{contadores["novo"]} alunos novos).',
                                'pauta'
                            )
                        else:
                            flash(
                                f'"{arquivo.filename}": {contadores["importado"]} importado(s), '
                                f'{contadores["novo"]} novo(s), '
                                f'{contadores["nome"]} nome(s) a confirmar.',
                                'pauta'
                            )
                    except Exception as e:
                        db.session.rollback()
                        mensagens_erro.append(f'"{arquivo.filename}": erro ao processar linhas — {str(e)}')
        except Exception as e:
            db.session.rollback()
            mensagens_erro.append(f'"{arquivo.filename}": erro ao guardar — {str(e)}')

    for msg in mensagens_erro:
        flash(msg, 'pauta')

    if total_processados:
        flash(f'{total_processados} ficheiro(s) Excel guardado(s) com sucesso.', 'pauta')
    elif not mensagens_erro:
        flash('Nenhum ficheiro Excel válido foi enviado.', 'pauta')

    return redirect(url_for('controle.central_verificacao'))

@controle_bp.route('/admin/grupos', methods=['POST'])
@login_required
def gerir_grupos():
    classe_numero = request.form.get('classe', '').strip()
    turma_nome = request.form.get('turma', '').strip()
    grupo_nome = request.form.get('grupo', '').strip()

    if not classe_numero or not turma_nome or not grupo_nome:
        flash('Classe, turma e grupo são obrigatórios.', 'grupo')
        return redirect(url_for('controle.central_verificacao'))

    try:
        classe_numero = int(classe_numero)
    except ValueError:
        flash('Classe inválida.', 'grupo')
        return redirect(url_for('controle.central_verificacao'))

    classe = Classe.query.filter_by(numero=classe_numero, deleted_at=None).first()
    if not classe:
        classe = Classe(numero=classe_numero, nome=f"{classe_numero}ª Classe")
        db.session.add(classe)
        db.session.flush()

    grupo = Grupo.query.filter_by(nome=grupo_nome, classe_id=classe.id, deleted_at=None).first()
    if not grupo:
        grupo = Grupo(nome=grupo_nome, classe_id=classe.id)
        db.session.add(grupo)
        db.session.flush()

    turma = Turma.query.filter_by(nome=turma_nome, grupo_id=grupo.id, deleted_at=None).first()
    if not turma:
        turma = Turma(nome=turma_nome, classe_id=classe.id, grupo_id=grupo.id)
        db.session.add(turma)

    try:
        db.session.commit()
        flash(f'Grupo "{grupo_nome}" e turma "{turma_nome}" criados/confirmados com sucesso!', 'grupo')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar grupo/turma: {str(e)}', 'grupo')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/carrossel', methods=['POST'])
@login_required
def atualizar_carrossel():
    posicoes_atualizadas = []

    for posicao in range(1, 6):
        nome_foto = _guardar_ficheiro(f'foto{posicao}')
        if not nome_foto:
            continue

        item = Carrossel.query.filter_by(ordem=posicao).first()
        if not item:
            item = Carrossel(imagem=nome_foto, ordem=posicao, ativo=True)
            db.session.add(item)
        else:
            item.imagem = nome_foto
            item.ativo = True

        posicoes_atualizadas.append(posicao)

    if not posicoes_atualizadas:
        flash('Nenhuma imagem foi selecionada — o carrossel não foi alterado.', 'carrossel')
        return redirect(url_for('controle.central_verificacao'))

    try:
        db.session.commit()
        posicoes_texto = ', '.join(str(p) for p in posicoes_atualizadas)
        flash(f'Carrossel atualizado com sucesso! Posição(ões) alterada(s): {posicoes_texto}.', 'carrossel')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar carrossel: {str(e)}', 'carrossel')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/substituir-pauta', methods=['POST'])
@login_required
def substituir_pauta():
    pendencia_id = request.form.get('id')
    if not pendencia_id:
        flash('Identificador de pendência inválido.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    pendencia = PendenciaPauta.query.get(pendencia_id)
    if not pendencia:
        flash('Registo de pendência não encontrado.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    try:
        notas_temp = NotaTemporaria.query.filter_by(pendencia_id=pendencia.id).all()
        for nt in notas_temp:
            # Substitui apenas a nota deste aluno + disciplina + período —
            # nunca as restantes notas da turma.
            Nota.query.filter_by(
                aluno_id=nt.aluno_id,
                disciplina=nt.disciplina,
                classe=nt.classe,
                turma=nt.turma,
                periodo=nt.periodo
            ).delete(synchronize_session=False)

            db.session.add(Nota(
                aluno_id=nt.aluno_id,
                disciplina=nt.disciplina,
                classe=nt.classe,
                turma=nt.turma,
                periodo=nt.periodo,
                nota_ac=nt.nota_ac,
                nota_pt=nt.nota_pt,
                nota_ap=nt.nota_ap,
                nota_exame=getattr(nt, "nota_exame", None),
            ))

        NotaTemporaria.query.filter_by(pendencia_id=pendencia.id).delete(synchronize_session=False)
        pendencia.status = 'resolvido'

        db.session.commit()
        flash('Pauta substituída com sucesso!', 'pendencia-sucesso')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar a substituição: {str(e)}', 'pendencia-erro')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/ignorar-pauta', methods=['POST'])
@login_required
def ignorar_pauta():
    pendencia_id = request.form.get('id')
    if not pendencia_id:
        flash('Identificador de pendência inválido.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    pendencia = PendenciaPauta.query.get(pendencia_id)
    if not pendencia:
        flash('Pendência não encontrada.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    try:
        NotaTemporaria.query.filter_by(pendencia_id=pendencia.id).delete(synchronize_session=False)
        pendencia.status = 'ignorado'

        db.session.commit()
        flash('A alteração foi descartada. A pauta original foi mantida.', 'pendencia-sucesso')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao ignorar pauta: {str(e)}', 'pendencia-erro')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/confirmar-aluno', methods=['POST'])
@login_required
def confirmar_aluno():
    pendencia_id = request.form.get('id')
    id_aluno = request.form.get('id_aluno')

    if not pendencia_id or not id_aluno:
        flash('Por favor, informe o ID correto do aluno.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    aluno_existe = Aluno.query.get(id_aluno)
    if not aluno_existe:
        flash(f'Aluno com ID {id_aluno} não foi encontrado na base de dados.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    pendencia = PendenciaPauta.query.get(pendencia_id)
    if not pendencia:
        flash('Pendência não encontrada.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    try:
        notas_temp = NotaTemporaria.query.filter_by(pendencia_id=pendencia.id).all()
        for nt in notas_temp:
            nova_nota = Nota(
                aluno_id=aluno_existe.id,
                disciplina=nt.disciplina,
                classe=nt.classe,
                turma=nt.turma,
                periodo=nt.periodo,
                nota_ac=nt.nota_ac,
                nota_pt=nt.nota_pt,
                nota_ap=nt.nota_ap,
                nota_exame=getattr(nt, "nota_exame", None),
            )
            db.session.add(nova_nota)

        NotaTemporaria.query.filter_by(pendencia_id=pendencia.id).delete(synchronize_session=False)
        pendencia.status = 'resolvido'

        db.session.commit()
        flash(f'Notas associadas com sucesso ao aluno ID {id_aluno}.', 'pendencia-sucesso')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao associar aluno: {str(e)}', 'pendencia-erro')

    return redirect(url_for('controle.central_verificacao'))


@controle_bp.route('/admin/adicionar-aluno', methods=['POST'])
@login_required
def adicionar_aluno():
    pendencia_id = request.form.get('id')
    if not pendencia_id:
        flash('Identificador de pendência inválido.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    pendencia = PendenciaPauta.query.get(pendencia_id)
    if not pendencia or not pendencia.nome_excel:
        flash('Não foi possível obter os dados do aluno a partir da pendência.', 'pendencia-erro')
        return redirect(url_for('controle.central_verificacao'))

    try:
        codigo = _gerar_codigo_estudante_unico()
        novo_aluno = Aluno(
            nome=pendencia.nome_excel,
            codigo_estudante=codigo,
        )
        db.session.add(novo_aluno)
        db.session.flush()

        # Tenta associar classe
        if pendencia.classe:
            classe_str = str(pendencia.classe).strip()
            classe_obj = None
            if classe_str.isdigit():
                classe_obj = Classe.query.filter_by(numero=int(classe_str), deleted_at=None).first()
            if not classe_obj:
                classe_obj = Classe.query.filter_by(nome=classe_str, deleted_at=None).first()
            if classe_obj:
                novo_aluno.classe_id = classe_obj.id

        notas_temp = NotaTemporaria.query.filter_by(pendencia_id=pendencia.id).all()
        for nt in notas_temp:
            nova_nota = Nota(
                aluno_id=novo_aluno.id,
                disciplina=nt.disciplina,
                classe=nt.classe,
                turma=nt.turma,
                periodo=nt.periodo,
                nota_ac=nt.nota_ac,
                nota_pt=nt.nota_pt,
                nota_ap=nt.nota_ap,
                nota_exame=getattr(nt, "nota_exame", None),
            )
            db.session.add(nova_nota)

        NotaTemporaria.query.filter_by(pendencia_id=pendencia.id).delete(synchronize_session=False)
        pendencia.status = 'resolvido'

        db.session.commit()
        flash(
            f'Aluno "{pendencia.nome_excel}" registado com sucesso! '
            f'ID: {codigo} | Senha: ESGAM000 | Notas importadas.',
            'pendencia-sucesso'
        )

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registar novo aluno: {str(e)}', 'pendencia-erro')

    return redirect(url_for('controle.central_verificacao'))


# ==========================================
# 5. SERVIÇOS DO PORTAL
# ==========================================

def carregar_portal(student_id):
    """Monta a vista do portal. Cálculos básicos vêm do model (media_parcial / media_com_exame)."""
    try:
        aluno = Aluno.query.filter_by(id=student_id, deleted_at=None).first()

        if not aluno:
            return {
                "aluno": {
                    "id": student_id, "nome": "—", "classe": "0",
                    "turma": "—", "grupo": "—", "codigo": None,
                    "media_geral": "—", "situacao": None,
                },
                "pauta_disciplinas": [],
                "aviso_painel": "Estudante não encontrado no sistema escolar."
            }

        aluno_view = {
            "id": aluno.id,
            "nome": aluno.nome,
            "classe": (aluno.classe_rel.numero if aluno.classe_rel else (aluno.classe_nome or "0")),
            "turma": aluno.turma_nome or "—",
            "grupo": aluno.grupo_nome or "—",
            "codigo": aluno.codigo_estudante or str(aluno.id),
            "media_geral": "—",
            "situacao": None,
        }

        notas = Nota.query.filter_by(aluno_id=student_id).all()

        from collections import defaultdict
        por_disciplina = defaultdict(lambda: {
            "nome": "",
            "t1_p1": None, "t1_p2": None, "t1_p3": None, "t1_p4": None, "mf1": None,
            "t2_p1": None, "t2_p2": None, "t2_p3": None, "t2_p4": None, "mf2": None,
            "t3_p1": None, "t3_p2": None, "t3_p3": None, "t3_p4": None, "mf3": None,
            "exame": None, "nota_global": None,
        })

        for n in notas:
            d = por_disciplina[n.disciplina]
            d["nome"] = n.disciplina
            periodo = (n.periodo or "").lower()
            mf = n.media_parcial  # propriedade do model

            if "1" in periodo or "primeiro" in periodo or "1º" in periodo:
                d["t1_p1"], d["t1_p2"], d["t1_p3"], d["mf1"] = n.nota_ac, n.nota_pt, n.nota_ap, mf
            elif "2" in periodo or "segundo" in periodo or "2º" in periodo:
                d["t2_p1"], d["t2_p2"], d["t2_p3"], d["mf2"] = n.nota_ac, n.nota_pt, n.nota_ap, mf
            elif "3" in periodo or "terceiro" in periodo or "3º" in periodo:
                d["t3_p1"], d["t3_p2"], d["t3_p3"], d["mf3"] = n.nota_ac, n.nota_pt, n.nota_ap, mf
            else:
                d["t1_p1"], d["t1_p2"], d["t1_p3"], d["mf1"] = n.nota_ac, n.nota_pt, n.nota_ap, mf

            if n.nota_exame is not None:
                d["exame"] = n.nota_exame

            mfs = [v for v in (d["mf1"], d["mf2"], d["mf3"]) if v is not None]
            if d["exame"] is not None and mfs:
                d["nota_global"] = round((sum(mfs) / len(mfs) + d["exame"]) / 2, 1)
            elif mfs:
                d["nota_global"] = round(sum(mfs) / len(mfs), 1)

        pauta_disciplinas = list(por_disciplina.values())

        todas = [d["nota_global"] for d in pauta_disciplinas if d.get("nota_global") is not None]
        if not todas:
            todas = []
            for d in pauta_disciplinas:
                for k in ("mf1", "mf2", "mf3"):
                    if d.get(k) is not None:
                        todas.append(d[k])
        if todas:
            aluno_view["media_geral"] = round(sum(todas) / len(todas), 1)
            aluno_view["situacao"] = "Aprovado" if aluno_view["media_geral"] >= 10 else "Reprovado"

        aviso_mensagem = None
        try:
            aviso = Aviso.query.filter_by(ativo=True).first()
            if aviso:
                aviso_mensagem = getattr(aviso, "mensagem", None) or getattr(aviso, "texto", None)
        except Exception:
            db.session.rollback()

        return {
            "aluno": aluno_view,
            "pauta_disciplinas": pauta_disciplinas,
            "aviso_painel": aviso_mensagem,
        }

    except Exception:
        logging.exception("Erro ao carregar o portal do estudante.")
        db.session.rollback()
        return {
            "aluno": {
                "id": student_id, "nome": "—", "classe": "0",
                "turma": "—", "grupo": "—", "codigo": None,
                "media_geral": "—", "situacao": None,
            },
            "pauta_disciplinas": [],
            "aviso_painel": "Erro ao carregar os dados.",
        }


# ==========================================
# 6. SERVIÇOS DE AUTENTICAÇÃO (LOGIN)
# ==========================================

def carregar_login():
    """Lógica preparatória para a página de login."""
    return {}