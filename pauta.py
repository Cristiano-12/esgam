from datetime import datetime
from io import BytesIO
from flask import Blueprint, render_template, send_file, abort
from models import ConfiguracaoSistema, PautaTurma

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None

pauta_bp = Blueprint("pauta", __name__)


def _celula_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def _titulo_amigavel(pauta, indice, linhas_preview=None):
    """Título legível — nunca o nome feio do ficheiro."""
    # Tenta ler título da própria pauta (linhas de cabeçalho da escola)
    if linhas_preview:
        for linha in linhas_preview[:4]:
            texto = " ".join(c.strip() for c in linha if c and str(c).strip())
            if "PAUTA" in texto.upper() or "CLASSE" in texto.upper():
                # encurta um pouco
                texto = " ".join(texto.split())
                if len(texto) > 80:
                    texto = texto[:77] + "…"
                return texto

    cat = (pauta.categoria or pauta.tipo or "").strip()
    data = pauta.data_publicacao.strftime("%d/%m/%Y") if pauta.data_publicacao else ""
    if cat and data:
        return f"{cat} · {data}"
    if cat:
        return cat
    if data:
        return f"Documento · {data}"
    return f"Documento {indice}"


def _extrair_folha_completa(conteudo_bytes, max_linhas=200):
    """
    Espelho da folha Excel: devolve todas as linhas como grelha
    (sem assumir que a 1ª linha é o único cabeçalho).
    """
    if not load_workbook or not conteudo_bytes:
        return []

    try:
        livro = load_workbook(BytesIO(conteudo_bytes), data_only=True, read_only=True)
        folha = livro.active
        brutas = list(folha.iter_rows(values_only=True))
        if not brutas:
            return []

        # Largura máxima real (ignora cauda vazia)
        max_col = 0
        for linha in brutas:
            for i, c in enumerate(linha):
                if c is not None and str(c).strip() != "":
                    max_col = max(max_col, i + 1)

        if max_col == 0:
            return []

        grelha = []
        for linha in brutas[:max_linhas]:
            vals = [_celula_str(c) for c in linha[:max_col]]
            while len(vals) < max_col:
                vals.append("")
            grelha.append(vals)

        # Remove linhas finais totalmente vazias
        while grelha and not any(c.strip() for c in grelha[-1]):
            grelha.pop()

        return grelha
    except Exception:
        return []


@pauta_bp.route("/consulta-turmas")
def consulta_turmas():
    config = ConfiguracaoSistema.query.first()
    modo_pauta_aberto = config.modo_pauta_aberto if config else True

    publicacoes = []
    pautas = (
        PautaTurma.query.filter_by(ativo=True)
        .order_by(PautaTurma.data_publicacao.desc())
        .all()
    )

    for i, pauta in enumerate(pautas, start=1):
        grelha = []
        mime = (pauta.mimetype or "").lower()
        nome = (pauta.arquivo or "").lower()
        e_excel = (
            "spreadsheet" in mime
            or "excel" in mime
            or nome.endswith((".xlsx", ".xls"))
            or (pauta.tipo or "").lower() in ("turmas", "notas")
            or bool(pauta.ficheiro)
        )

        if e_excel and pauta.ficheiro:
            grelha = _extrair_folha_completa(pauta.ficheiro)

        titulo = _titulo_amigavel(pauta, i, grelha)

        publicacoes.append({
            "id": pauta.id,
            "titulo": titulo,
            "descricao": (pauta.categoria or pauta.tipo or "Documento").strip(),
            "data_publicacao": pauta.data_publicacao,
            "grelha": grelha,
            "tem_preview": bool(grelha),
            "total_linhas": len(grelha),
            "mime": pauta.mimetype or "",
        })

    return render_template(
        "pauta.html",
        publicacoes=publicacoes,
        modo_pauta_aberto=modo_pauta_aberto,
    )


@pauta_bp.route("/consulta-turmas/ficheiro/<int:pauta_id>")
def descarregar_ficheiro(pauta_id):
    pauta = PautaTurma.query.get_or_404(pauta_id)
    if not pauta.ficheiro:
        abort(404)

    nome = pauta.arquivo or f"documento_{pauta.id}.xlsx"
    return send_file(
        BytesIO(pauta.ficheiro),
        mimetype=pauta.mimetype
        or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome,
    )
